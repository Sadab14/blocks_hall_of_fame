import os
from flask import Flask, render_template, redirect
from openpyxl import load_workbook

app = Flask(__name__)

MINIFIG_THRESHOLDS = [
    (250, "Grand Master"),
    (150, "Platinum"),
    (100, "Gold"),
    (50, "Silver"),
    (20, "Bronze"),
    (0, "Beginner")
]
POINTS_THRESHOLDS = [
    (500, "Grand Master"),
    (375, "Platinum"),
    (250, "Gold"),
    (100, "Silver"),
    (50, "Bronze"),
    (0, "Beginner")
]

def get_tier(value, thresholds):
    for thresh, name in thresholds:
        if value >= thresh:
            return name
    return thresholds[-1][1]

def load_data():
    path = "hall_of_fame.xlsx"
    if not os.path.exists(path):
        return []

    wb = load_workbook(path, data_only=True)
    ws = wb.active

    # Try to detect header positions, fall back to defaults
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    def idx(name, fallback):
        try:
            return headers.index(name)
        except ValueError:
            return fallback

    # common header names; adjust if your sheet differs
    i_rank = idx("Rank", 0)
    i_profile = idx("ProfileImage", 1)
    i_name = idx("CollectorName", 2)
    i_nick = idx("Nickname", 3)
    i_minifigs = idx("Total_Minifigs", 4)
    i_points = idx("Total_Points", 5)
    i_title = idx("Special_Title", 6)

    tier_order = ["Beginner", "Bronze", "Silver", "Gold", "Platinum", "Grand Master"]
    rank_map = {t: i for i, t in enumerate(tier_order)}

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        def val(i):
            try:
                v = row[i]
            except Exception:
                v = None
            return v if v is not None else "—"

        try:
            minifigs = int(row[i_minifigs] or 0)
        except Exception:
            minifigs = 0
        try:
            points = int(row[i_points] or 0)
        except Exception:
            points = 0

        points_tier = get_tier(points, POINTS_THRESHOLDS)
        minifig_tier = get_tier(minifigs, MINIFIG_THRESHOLDS)
        overall = points_tier if rank_map[points_tier] >= rank_map[minifig_tier] else minifig_tier
        source = "Points" if rank_map[points_tier] >= rank_map[minifig_tier] else "Minifigs"

        record = {
            "Rank": val(i_rank),
            "ProfileImage": val(i_profile),
            "CollectorName": val(i_name),
            "Nickname": val(i_nick),
            "Total_Minifigs": minifigs,
            "Total_Points": points,
            "Special_Title": val(i_title),
            "Points_Tier": points_tier,
            "Minifig_Tier": minifig_tier,
            "Overall_Tier": overall,
            "Tier_Source": source
        }
        data.append(record)

    return data

@app.route('/')
def home():
    data = sorted(load_data(), key=lambda x: x["Total_Points"], reverse=True)
    top3 = data[:3]
    return render_template("home.html", top3=top3)

@app.route('/top-collectors')
def top_collectors():
    data = sorted(load_data(), key=lambda x: x["Total_Points"], reverse=True)
    return render_template("top_collectors.html", collectors=data)

@app.route('/submit')
def submit():
    return redirect("https://forms.gle/5wYMhzYjZS9vY9Sv8")

@app.route('/about')
def about():
    return render_template('about.html')

@app.route('/sheet')
def sheet():
    try:
        return render_template('Sheet3.html')
    except Exception as e:
        return f"Error loading sheet: {str(e)}", 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
